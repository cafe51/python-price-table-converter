from tabulate import tabulate
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

def sudeste(preço): 
  return preço * 60/100


def amazonas(preço): 
  return preço * 100/100


def frete(preço):
  return preço * 15/100


def final_price_with_frete(estado: str, preço_inicial: float):
  preço_final = (preço_inicial + estado(preço_inicial))
  return round(preço_final + frete(preço_final), 2)


def print_changes(estado, table):
  local = 'Amazonas' if estado == amazonas else 'Sudeste'
  header = ['Preço de Compra', f'Preço de Venda para o {local}']
  print(tabulate(table, headers=header, tablefmt='fancy_grid', stralign='center'))


def change_header_color(estado, worksheet):
  color = '24BF0F' if estado == amazonas else '57B8F0'
  fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
  for col in worksheet.iter_cols(min_row=1, max_row=1):
    for cell in col:
        cell.fill = fill

def change_values(estado, worksheet, table):
  for index in range(2, len(worksheet['E'])):
    old_value = worksheet['E'][index].value
    new_value = final_price_with_frete(estado, worksheet['E'][index].value)
    worksheet['E'][index].value = new_value
    table.append([f'R$ {old_value}', f'R$ {new_value}'])


def change_table(estado, excel_file):
  workbook = load_workbook(excel_file)
  worksheet = workbook['Table 1']
  table = []

  change_header_color(estado, worksheet)
  change_values(estado, worksheet, table)

  local = 'Amazonas' if estado == amazonas else 'Sudeste'
  file_name = f'/home/japhe/Downloads/tabeladiadasmaes2023/tabela_de_bolsas_{local}.xlsx'
  workbook.save(file_name)

  print_changes(estado, table)

def exec(excel_file):
  while True:
    local = input('''
    Escolha uma tabela
    1 - Amazonas
    2 - Sudeste
    3 - As duas
    ''')

    if local == '1':
      change_table(amazonas, excel_file)
      break
    elif local == '2':
      change_table(sudeste, excel_file)
      break
    elif local == '3':
      change_table(sudeste, excel_file)
      change_table(amazonas, excel_file)
      break
    else:
      print('RESPOSTA ERRADA TENTE DE NOVO')

diretorio = '/home/japhe/Downloads/tabeladiadasmaes2023/'

arquivos_excel = [arquivo for arquivo in os.listdir(diretorio) if arquivo.endswith('.xlsx')]

excel_file = ''

if len(arquivos_excel) == 0:
  print('Não há arquivos nessa pasta')
elif len(arquivos_excel) == 1:
  excel_file = arquivos_excel[0]
  exec(diretorio + excel_file)
elif len(arquivos_excel) > 1:
  options = ''
  for arquivo_index in range(len(arquivos_excel)):
    options += f'''
    {arquivo_index + 1} - { arquivos_excel[arquivo_index]} 
    '''
  file_index = input(f'''
  Escolha uma tabela:
  {options}
  ''')
  if (not file_index.isdigit() or (int(file_index) - 1) not in range(len(arquivos_excel))):
    print('ESCOLHA UMA OPÇÃO VÁLIDA')
  else:
    excel_file = arquivos_excel[int(file_index) - 1]
    exec(diretorio + excel_file)
  # excel_file = arquivos_excel[file_index]
  # print(excel_file)



# for arquivo in arquivos_excel:
#     print(arquivo)

